import pandas as pd
import getpass
from openpyxl import Workbook

testList = [
    {'idx': 1, 'title': '케라시스 프로폴리스 샴푸/트리트먼트 1L x1개', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=31621637&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=1&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창  열기</button'},
    {'idx': 2, 'title': '케라시스 퍼퓸 체리블라썸 샴푸/린스 1L x2개 + 샴푸 180ml 증정', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=37522614&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=2&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 3, 'title': '케라시스 어드밴스드 샴푸/트리트먼트 600ml x2개 + 샴푸 180ml', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=33176633&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=3&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 4, 'title': '[로레알파리] NEW 헤어오일 런칭기념 전상품 20%쿠폰 ', 'status': '바로구매',
        'link': 'http://www.gsshop.com/deal/deal.gs?dealNo=1032822233&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=4&amp;lseq=405305\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 5, 'title': '려 흑운 모근강화 앤 볼륨케어 컨디셔너 550ml 2개', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=97561546&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=5&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 6, 'title': '[로레알] [2개] 드림랭스 컨디셔닝 헤어팩 410ml', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=1022592101&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=6&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 7, 'title': '케라시스 샤이닝 데미지 린스 4000ml + 샴푸 180ml 증정', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=31648696&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=7&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 8, 'title': '[착한팩토리] 밀크단백질90% 트리트먼트 1,000ml', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=19919459&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=8&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 9, 'title': '[바론] 모링가 체리블라썸 샴푸 500g x2', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=32269545&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=9&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 10, 'title': '[증정] 휘엔느 모링가 단백질 대용량 저자극 트리트먼트 1000ml', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=30341765&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=10&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 11, 'title': '[TV상품]테라픽 프래그런스 헤어팩 500ml', 'status': '바로구매', 'link': 'http://www.gsshop.com/prd/prd.gs?prdid=57858930&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=11&amp;lseq=405304\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'}, {
        'idx': 12, 'title': '[로레알] 히알루론산 컨디셔닝 헤어팩 440ml', 'status': '바로구매', 'link': 'http://www.gsshop.com/prd/prd.gs?prdid=1031220150&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=12&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 13, 'title': '아윤채 리밸런싱 마스크 200ml', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=93213699&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=13&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 14, 'title': '[러쉬]수퍼 밀크 100g - 컨디셔닝 헤어 프라이머', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=66760546&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=14&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 15, 'title': '미친스킨 맥주효모 헤어팩 200g 2개', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=57023093&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=15&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 16, 'title': '도브 너리싱 오일 케어 컨디셔너 660ml 2개', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=93545626&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=16&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 17, 'title': '아윤채 인핸싱 실키 마스크 200ml', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=93213687&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=17&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'},
    {'idx': 18, 'title': '케라시스 엑스트라 데미지 케어 린스 750ml', 'status': '바로구매',
        'link': 'http://www.gsshop.com/prd/prd.gs?prdid=27385168&amp;sectid=1661646&amp;lsectid=1660624&amp;msectid=1661646&amp;rank=18&amp;lseq=405302\'); return false;"><i class="sprite-new-tab"></i>새창 열기</button'}
]


def makeExcel(category, dataList):
    print("MakeExcel Start")
    wb = Workbook()
    ws = wb.active
    ws.title = "dataList"  # 엑셀 시트명 변경
    ws['A1'] = category  # 한 행 입력
    ws.append(['idx', 'title', 'status', 'link'])
    user = getpass.getuser()

    for dict in dataList:
        ws.append([dict['idx'], dict['title'], dict['status'], dict['link']])

    filename = f'C:\\Users\\{user}\\Desktop\\productScrap.xlsx'
    wb.save(filename)
    print("MakeExcel End")


# makeExcel('린스/컨디셔너', testList)


#     # 엑셀파일 쓰기
# write_wb = Workbook()

# # 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')

# # Sheet1에다 입력
# write_ws = write_wb.active
# write_ws['A1'] = '숫자'

# #행 단위로 추가
# write_ws.append([1,2,3])

# #셀 단위로 추가
# write_ws.cell(5, 5, '5행5열')
# write_wb.save("C:/Users/Administrator/Desktop/기준/프로그래밍/과제대행/주식데이터크롤링/숫자.xlsx")

# raw_data = {
#     'idx': [dict.get('idx', 'null')],
#     'title': [dict.get('title', 'null')],
#     'status': [dict.get('status', 'null')],
#     'link': [dict.get('link', 'null')]
# }  # 리스트 자료형으로 생성
# raw_data1 = pd.DataFrame(raw_data)  # 데이터 프레임으로 전환 및 생성
# raw_data2 = pd.DataFrame(raw_data)  # 데이터 프레임으로 전환 및 생성
# xlxs_dir = 'sample.xlsx'  # 경로 및 파일명 설정
# with pd.ExcelWriter(xlxs_dir) as writer:
#     # raw_data1 시트에 저장
#     raw_data1.to_excel(writer, sheet_name='raw_data1')
#     # raw_data2 시트에 저장
#     raw_data2.to_excel(writer, sheet_name='raw_data2')
